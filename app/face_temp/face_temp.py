from flask import render_template, request, redirect, url_for, session
from . import face_temp_bp
from flask import Blueprint, render_template, request, redirect, url_for, Response, jsonify, session, make_response
import mysql.connector
import cv2
from PIL import Image
import numpy as np
import os
import time
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import date
import serial
# from flask_login import login_required


cnt = 0
pause_cnt = 0
justscanned = False

# date and time
def datetoday2():
    return date.today().strftime("%d-%B-%Y")

# Database connection with xampp
mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    passwd="",
    database="facesence_db"
)
mycursor = mydb.cursor()

# files for algorithm of face detection
face_cascade = cv2.CascadeClassifier(
    r"C:\Users\James Reniel\mainProgram\WebApp\FaceSenseTempTrack\final-thesis-project\face-recognition-flask-attendance\app\face_temp\resources\haarcascade_frontalface_default.xml")

sample_cam = 0 


# Generate dataset
def generate_dataset(nbr):
    face_classifier = cv2.CascadeClassifier(
        r"C:\Users\James Reniel\mainProgram\WebApp\FaceSenseTempTrack\final-thesis-project\face-recognition-flask-attendance\app\face_temp\resources\haarcascade_frontalface_default.xml")

    def face_cropped(img):
        faces = face_classifier.detectMultiScale(img, 1.3, 5)

        if len(faces) == 0:
            return None

        x, y, w, h = faces[0]
        cropped_face = img[y:y + h, x:x + w]

        return cropped_face

    cap = cv2.VideoCapture(sample_cam)  # Update with the correct camera index or video path

    mycursor.execute("SELECT IFNULL(MAX(img_id), 0) FROM img_dataset")
    row = mycursor.fetchone()
    last_id = row[0]

    img_id = last_id
    max_img_id = img_id + 100
    count_img = 0

    while True:
        ret, img = cap.read()

        if not ret:
            break

        cropped_img = face_cropped(img)

        if cropped_img is not None:
            count_img += 1
            img_id += 1
            face_colored = cv2.resize(cropped_img, (200, 200))
            
            # Convert the colored face to grayscale
            face_gray = cv2.cvtColor(face_colored, cv2.COLOR_BGR2GRAY)

            file_name_path = "dataset/" + nbr + "." + str(img_id) + ".jpg"
            cv2.imwrite(file_name_path, face_gray)  # Saving the grayscale image
            
            cv2.putText(face_gray, str(count_img), (40, 25), cv2.FONT_HERSHEY_COMPLEX, 1, (0, 255, 0), 2)
            
            mycursor.execute("INSERT INTO img_dataset (img_id, img_person) VALUES (%s, %s)", (img_id, nbr))
            mydb.commit()
            
            frame = cv2.imencode('.jpg', face_gray)[1].tobytes()
            yield (b'--frame\r\n' b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')
            
            if cv2.waitKey(1) == 13 or int(img_id) == int(max_img_id):
                break


                cap.release()
                cv2.destroyAllWindows()


@face_temp_bp.route('/train_classifier/<nbr>')
def train_classifier(nbr):
    dataset_dir = r"C:\\Users\\James Reniel\\mainProgram\\WebApp\\FaceSenseTempTrack\\final-thesis-project\\face-recognition-flask-attendance\\dataset"

    path = [os.path.join(dataset_dir, f) for f in os.listdir(dataset_dir)]
    faces = []
    ids = []

    for image in path:
        img = Image.open(image).convert('L')
        imageNp = np.array(img, 'uint8')
        id = int(os.path.split(image)[1].split(".")[1])

        faces.append(imageNp)
        ids.append(id)
    ids = np.array(ids)

    # Check if there are enough samples to train the classifier
    if len(set(ids)) < 2:
        return "Error: Insufficient training data. You need at least two samples to learn a model."

    # Train the classifier and save
    clf = cv2.face.LBPHFaceRecognizer_create()
    try:
        clf.train(faces, ids)
        clf.write("classifier.xml")
        return redirect(url_for('face_temp.index'))
    except cv2.error as e:
        return f"Error: An error occurred during training: {str(e)}"

# Face Recognition
def face_recognition():  # generate frame by frame from camera
    # ser = serial.Serial('COM9', baudrate=9600, timeout=1)
    def draw_boundary(img, classifier, scaleFactor, minNeighbors, color, text, clf):
        gray_image = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        features = classifier.detectMultiScale(gray_image, 1.1, minNeighbors)
        

        global justscanned
        global pause_cnt

        pause_cnt += 1

        coords = []

        for (x, y, w, h) in features:
            # arduinoData = ser.readline().strip()
            # tempvalue = arduinoData.decode('ascii')
            tempvalue = "Variable"
            # print("Received temperature:", tempvalue)
 
            # Calculate the coordinates for the bottom rectangle and centered text
            rect_height = 30  # Height of the rectangle for the text
            rect_top_left = (x, y + h - rect_height)
            rect_bottom_right = (x + w, y + h)
            text_size = cv2.getTextSize( tempvalue + " C", cv2.FONT_HERSHEY_SIMPLEX, 0.8, 2)[0]  # Get text size
            text_x = x + (w - text_size[0]) // 2  # Center text horizontally
            text_y = y + h - (rect_height - text_size[1]) // 2  # Center text vertically

            # Draw the rectangle at the bottom of the detected face region
            cv2.rectangle(img, rect_top_left, rect_bottom_right, (153, 255, 255), cv2.FILLED)

            # Draw the centered text inside the rectangle (Variable = tempvalue) 
            cv2.putText(img, f"{tempvalue} C", (text_x, text_y), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 0, 0), 2)
            
            cv2.rectangle(img, (x, y), (x + w, y + h), color, 2)
            id, pred = clf.predict(gray_image[y:y + h, x:x + w])
            confidence = int(100 * (1 - pred / 300))

            if confidence > 80 and not justscanned:

                global cnt
                cnt += 1

                n = (100 / 30) * cnt
                # w_filled = (n / 100) * w
                w_filled = (cnt / 30) * w

                cv2.putText(img, str(int(n)) + ' %', (x + 20, y + h + 28), cv2.FONT_HERSHEY_SIMPLEX, 0.8,
                            (153, 255, 255), 2, cv2.LINE_AA)

                cv2.rectangle(img, (x, y + h + 40), (x + w, y + h + 50), color, 2)
                cv2.rectangle(img, (x, y + h + 40), (x + int(w_filled), y + h + 50), (153, 255, 255), cv2.FILLED)

                mycursor.execute("select a.img_person, b.prs_name, b.prs_id, b.prs_course "
                                 "  from img_dataset a "
                                 "  left join prs_mstr b on a.img_person = b.prs_nbr "
                                 " where img_id = " + str(id))
                row = mycursor.fetchone()


                if row is not None:
                    try:
                        pnbr = row[0]
                        pname = row[1]
                        pid = row[2]
                        pcourse = row[3] 
                         # Corrected the index to 3 for the course
                    except IndexError:
                        pnbr = "N/A"
                        pname = "N/A"
                        pid = "N/A"
                        pcourse = "N/A"
                else:
                    pnbr = "N/A"
                    pname = "N/A"
                    pid = "N/A"
                    pcourse = "N/A"


                if int(cnt) == 30:
                    cnt = 0

                    mycursor.execute("insert into accs_hist (accs_date, accs_prsn) values('" + str(
                        date.today()) + "', '" + pnbr + "')")
                    mydb.commit()

                    cv2.putText(img, pname + ' | ' + pid, (x - 10, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.8,
                                (153, 255, 255), 2, cv2.LINE_AA)
                    time.sleep(1)

                    justscanned = True
                    pause_cnt = 0

            else:
                if not justscanned:
                    cv2.putText(img, 'UNKNOWN', (x, y - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2, cv2.LINE_AA)
                else:
                    cv2.putText(img, ' ', (x, y - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2, cv2.LINE_AA)

                if pause_cnt > 80:
                    justscanned = False

            coords = [x, y, w, h]
        return coords

    def recognize(img, clf, faceCascade):
        coords = draw_boundary(img, faceCascade, 1.1, 10, (255, 255, 0), "Face", clf)
        return img

    faceCascade = cv2.CascadeClassifier(
        r"C:\Users\James Reniel\mainProgram\WebApp\FaceSenseTempTrack\final-thesis-project\face-recognition-flask-attendance\app\face_temp\resources\haarcascade_frontalface_default.xml")
    clf = cv2.face.LBPHFaceRecognizer_create()
    clf.read("classifier.xml")

    wCam, hCam = 400, 400

    cap = cv2.VideoCapture(sample_cam)
    cap.set(3, wCam)
    cap.set(4, hCam)

    while True:
        ret, img = cap.read()
        img = recognize(img, clf, faceCascade)

        frame = cv2.imencode('.jpg', img)[1].tobytes()
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n\r\n')

        key = cv2.waitKey(1)
        if key == 27:
            break



 
# def faceDetection():  # generate frame by frame from camera
#     cap = cv2.VideoCapture(0)  # use 0 for web camera
#     ser = serial.Serial('COM9', baudrate=9600, timeout=1)
 
#     while True:
#         success, img = cap.read()
 
#         gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
 
#         faces = face_cascade.detectMultiScale(gray, 1.1, 4)
 
#         for (x, y, w, h) in faces:
#             cv2.rectangle(img, (x, y), (x + w, y + h), (255, 0, 0), 2)
 
#             arduinoData = ser.readline().strip()
#             tempvalue = arduinoData.decode('ascii')
#             print("Received temperature:", tempvalue)
 
#             cv2.putText(img, tempvalue + ' C', (x + w - 70, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (153, 255, 255), 2, cv2.LINE_AA)
 
#         frame = cv2.imencode('.jpg', img)[1].tobytes()
#         yield (b'--frame\r\n'b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')
 
#         key = cv2.waitKey(1)
#         if key == 27:
#             break
 
 
# @face_temp_bp.route('/temp')
# def temp():
#     # Video streaming route. Put this in the src attribute of an img tag
#     return Response(faceDetection(), mimetype='multipart/x-mixed-replace; boundary=frame')
 


@face_temp_bp.route('/video_feed')
def video_feed():
    # Video streaming route. Put this in the src attribute of an img tag
    return Response(face_recognition(), mimetype='multipart/x-mixed-replace; boundary=frame')




@face_temp_bp.route('/', methods = ['GET', 'POST'])
def index():
    mycursor.execute("select prs_nbr, prs_name, prs_id, prs_course, prs_active, prs_added from prs_mstr")
    data = mycursor.fetchall()
    return render_template('index.html', data=data)


# def video_feed_temp():
#     # Video streaming route. Put this in the src attribute of an img tag
#     return Response(faceDetection(), mimetype='multipart/x-mixed-replace; boundary=frame')

# Add student
@face_temp_bp.route('/addprsn')
def addprsn():
    mycursor.execute("select ifnull(max(prs_nbr) + 1, 101) from prs_mstr")
    row = mycursor.fetchone()
    nbr = row[0]
    # print(int(nbr))

    return render_template('addprsn.html', newnbr=int(nbr))


# Submit to add student
@face_temp_bp.route('/addprsn_submit', methods=['POST'])
def addprsn_submit():
    prsnbr = request.form.get('txtnbr')
    prsname = request.form.get('txtname').title()
    prsid = request.form.get('txtid').upper()
    prscourse = request.form.get('optcourse')

    mycursor.execute("""INSERT INTO `prs_mstr` (`prs_nbr`, `prs_name`, `prs_id`, `prs_course`) VALUES
                    ('{}', '{}', '{}', '{}')""".format(prsnbr, prsname, prsid, prscourse))
    mydb.commit()

    # return redirect(url_for('home'))
    return redirect(url_for('face_temp.vfdataset_page', prs=prsnbr))


# Route for classifier
@face_temp_bp.route('/vfdataset_page/<prs>')
def vfdataset_page(prs):
    return render_template('gendataset.html', prs=prs)


#
@face_temp_bp.route('/vidfeed_dataset/<nbr>')
def vidfeed_dataset(nbr):
    # Video streaming route. Put this in the src attribute of an img tag
    return Response(generate_dataset(nbr), mimetype='multipart/x-mixed-replace; boundary=frame')


#
@face_temp_bp.route('/fr_page')
def fr_page():
    """Video streaming home page."""
    mycursor.execute("select a.accs_id, a.accs_prsn, b.prs_name, b.prs_id, b.prs_course, a.accs_added "
                     "  from accs_hist a "
                     "  left join prs_mstr b on a.accs_prsn = b.prs_nbr "
                     " where a.accs_date = curdate() "
                     " order by 1 desc")
    data = mycursor.fetchall()

    return render_template('fr_page.html', data=data)


#
@face_temp_bp.route('/countTodayScan')
def countTodayScan():
    mydb = mysql.connector.connect(
        host="localhost",
        user="root",
        passwd="",
        database="facesence_db"
    )
    mycursor = mydb.cursor()

    mycursor.execute("select count(*) "
                     "  from accs_hist "
                     " where accs_date = curdate() ")
    row = mycursor.fetchone()
    rowcount = row[0]

    return jsonify({'rowcount': rowcount})


#
@face_temp_bp.route('/loadData', methods=['GET', 'POST'])
def loadData():
    mydb = mysql.connector.connect(
        host="localhost",
        user="root",
        passwd="",
        database="facesence_db"
    )
    mycursor = mydb.cursor()

    mycursor.execute(
        "select a.accs_id, b.prs_name, b.prs_id, b.prs_course, date_format(a.accs_added, '%H:%i:%s') "
        "  from accs_hist a "
        "  left join prs_mstr b on a.accs_prsn = b.prs_nbr "
        " where a.accs_date = curdate() "
        " order by 1 desc")
    data = mycursor.fetchall()

    return jsonify(response=data)

@face_temp_bp.route('/delete_student/<string:prs_nbr>', methods=['GET', 'POST'])
def delete_student(prs_nbr):
    if request.method == 'POST':
        # Perform deletion logic here

        # Check if images exist in the dataset before deletion
        mycursor.execute("SELECT COUNT(*) FROM img_dataset WHERE img_person = %s", (prs_nbr,))
        row = mycursor.fetchone()
        num_images = row[0]

        if num_images > 0:
            # Delete the student from prs_mstr and img_dataset
            mycursor.execute("DELETE FROM prs_mstr WHERE prs_nbr = %s", (prs_nbr,))
            mycursor.execute("DELETE FROM img_dataset WHERE img_person = %s", (prs_nbr,))
            mydb.commit()

            # Additional code to delete images from the dataset directory
            dataset_dir = r"C:\\Users\\James Reniel\\mainProgram\\WebApp\\FaceSenseTempTrack\\final-thesis-project\\face-recognition-flask-attendancee\\dataset" 
            images_to_delete = [os.path.join(dataset_dir, f) for f in os.listdir(dataset_dir) if f.startswith(prs_nbr)]
            for image_path in images_to_delete:
                os.remove(image_path)

            return redirect(url_for('face_temp.index'))
        else:
            return "No images found for this student. Deletion aborted."
    else:
        return render_template('index.html', prs_nbr=prs_nbr)

@face_temp_bp.route('/update_student/<int:student_id>', methods=['POST'])
def update_student(student_id):
    if request.method == 'POST':
        # Get the updated data from the form
        new_name = request.form.get('newName').title()
        new_id = request.form.get('newID').upper()
        new_course = request.form.get('newCourse')

        # Perform the update logic in the database
        mycursor.execute("UPDATE prs_mstr SET prs_name = %s, prs_id = %s, prs_course = %s WHERE prs_nbr = %s",
                         (new_name, new_id, new_course, student_id))
        mydb.commit()

        return redirect(url_for('face_temp.index'))  # Redirect to the main page after updating

    return render_template('index.html', student_id=student_id)


# Generated excel files
@face_temp_bp.route('/downloadExcel')
def download():
    mycursor.execute("SELECT prs_name, prs_id, prs_course, prs_added FROM prs_mstr")
    data = mycursor.fetchall()

    # Sort data based on the first letter of the student name
    data.sort(key=lambda x: x[0][0])  # Assuming the student name is in the second column (index 1)

    # Generate Excel file
    df = pd.DataFrame(data, columns=['Student Name', 'Student ID', 'Course', 'Time and Date'])

    # Create Excel writer
    excel_file = BytesIO()
    writer = pd.ExcelWriter(excel_file, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Sheet1')

    # Get the workbook and sheet
    writer.book.save(excel_file)
    writer.book.close()
    excel_file.seek(0)
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook['Sheet1']

    # Apply color design to header column
    header_fill = PatternFill(start_color="ADD8E6FF", end_color="ADD8E6FF", fill_type="solid")
    for cell in sheet[1]:
        cell.fill = header_fill

    # Customize column width
    column_widths = [30, 30, 20, 40]  # Specify the width of each column
    for i, width in enumerate(column_widths, start=1):
        column_letter = openpyxl.utils.get_column_letter(i)
        sheet.column_dimensions[column_letter].width = width

    # Customize row height
    row_height = 25  # Specify the height of each row
    for row in sheet.iter_rows():
        for cell in row:
            sheet.row_dimensions[cell.row].height = row_height

    # Customize font theme and size
    font_theme = "Calibri"  # Specify the font theme
    font_size = 14  # Specify the font size

    font = Font(name=font_theme, size=font_size)
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = font

    # Center-align and middle-align cells
    alignment = Alignment(horizontal='center', vertical='center')
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = alignment

    # Save the modified workbook to the Excel file buffer
    excel_file.seek(0)
    workbook.save(excel_file)
    workbook.close()

    # Create a response with the Excel file
    response = make_response(excel_file.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=Data of student.xlsx'

    return response


# Generate a downloadable Excel file
@face_temp_bp.route('/download_data')
def download_data():
    mydb = mysql.connector.connect(
        host="localhost",
        user="root",
        passwd="",
        database="facesence_db"
    )
    mycursor = mydb.cursor()

    mycursor.execute("SELECT b.prs_name, b.prs_id, b.prs_course, DATE_FORMAT(a.accs_added, '%H:%i:%s') "
                     "FROM accs_hist a "
                     "LEFT JOIN prs_mstr b ON a.accs_prsn = b.prs_nbr "
                     "WHERE a.accs_date = CURDATE() "
                     "ORDER BY 1 DESC")
    data = mycursor.fetchall()

    # Sort data based on the first letter of the student name
    data.sort(key=lambda x: x[0][0])  # Assuming the student name is in the first column (index 0)

    df = pd.DataFrame(data, columns=['Name', 'Student ID', 'Course', 'Time'])

    # Create Excel workbook and sheet
    excel_file = BytesIO()
    writer = pd.ExcelWriter(excel_file, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Sheet1')

    # Get the workbook and sheet
    writer.book.save(excel_file)
    writer.book.close()
    excel_file.seek(0)
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook['Sheet1']

    # Apply color design to header column
    header_fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid")
    for cell in sheet[1]:
        cell.fill = header_fill

    # Customize column width
    column_widths = [35, 20, 20, 20]  # Specify the width of each column
    for i, width in enumerate(column_widths, start=1):
        column_letter = openpyxl.utils.get_column_letter(i)
        sheet.column_dimensions[column_letter].width = width

    # Customize row height
    row_height = 25  # Specify the height of each row
    for row in sheet.iter_rows():
        for cell in row:
            sheet.row_dimensions[cell.row].height = row_height

    # Customize font theme and size
    font_theme = "Calibri"  # Specify the font theme
    font_size = 14  # Specify the font size

    font = Font(name=font_theme, size=font_size)
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = font

    # Center-align and middle-align cells
    alignment = Alignment(horizontal='center', vertical='center')
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = alignment

    # Save the modified workbook to the Excel file buffer
    excel_file.seek(0)
    workbook.save(excel_file)
    workbook.close()

    # Create a response with the Excel file
    response = make_response(excel_file.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={datetoday2()} - today_scan_data.xlsx'

    return response


# Delete the scanned data
@face_temp_bp.route('/delete_all_data', methods=['POST'])
def delete_all_data():
    mydb = mysql.connector.connect(
        host="localhost",
        user="root",
        passwd="",
        database="facesence_db"
    )
    mycursor = mydb.cursor()

    mycursor.execute("DELETE FROM accs_hist WHERE accs_date = CURDATE()")

    mydb.commit()
    mycursor.close()
    mydb.close()

    return "Data deleted successfully"


# Delete the scanned data
@face_temp_bp.route('/delete_data', methods=['POST'])
def delete_data():
    mydb = mysql.connector.connect(
        host="localhost",
        user="root",
        passwd="",
        database="facesence_db"
    )
    mycursor = mydb.cursor()

    accs_id = request.form.get('id')  # Get the ID of the row to delete

    mycursor.execute("DELETE FROM accs_hist WHERE accs_id = %s", (accs_id,))

    mydb.commit()
    mycursor.close()
    mydb.close()

    return "Data deleted successfully"

